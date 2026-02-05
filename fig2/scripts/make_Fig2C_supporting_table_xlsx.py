#!/usr/bin/env python3
import argparse
import re
from datetime import datetime
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

def short_name(x: str) -> str:
    x = str(x)
    m = re.match(r'^(CNE\d+|mm\d+)', x)
    if m:
        return m.group(1)
    return x.split("_")[0]

def guess_name_col(df):
    for c in ["enhancer", "Enhancer", "name", "Name", "id", "ID"]:
        if c in df.columns:
            return c
    return df.columns[0]

def guess_loc_col(df):
    for c in ["location","Location","group","Group","type","Type","class","Class","region","Region","category","Category"]:
        if c in df.columns:
            return c
    return None

def autosize_columns(ws, min_w=10, max_w=60):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_w, min(max_w, max_len + 2))

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--tsv", required=True, help="Input TSV for Fig2C (mean signals).")
    ap.add_argument("--out", required=True, help="Output .xlsx path.")
    ap.add_argument("--title", default="Supporting Data for Fig. 2C — mean ChIP-seq signal across Gli3 limb enhancers (mm9)")
    ap.add_argument("--subtitle", default="WT hindlimb E11.5 (mm9). Signals are mean bigWig values across each enhancer interval.")
    args = ap.parse_args()

    df = pd.read_csv(args.tsv, sep="\t")
    if df.empty:
        raise SystemExit(f"TSV is empty: {args.tsv}")

    name_col = guess_name_col(df)
    loc_col = guess_loc_col(df)

    # Required columns (your pipeline)
    req = ["H3K4me1_mean", "H3K27ac_mean", "H3K4me3_mean"]
    missing = [c for c in req if c not in df.columns]
    if missing:
        raise SystemExit(f"Missing required columns: {missing}\nFound columns: {list(df.columns)}")

    # Clean numeric
    for c in req:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)

    # Location
    if loc_col is None:
        df["Location"] = "unknown"
    else:
        df["Location"] = df[loc_col].astype(str).str.strip().str.lower()
        df["Location"] = df["Location"].replace({"intron":"intronic", "intronic":"intronic", "up":"upstream", "upstream":"upstream"})

    # Add display columns
    df["Enhancer_ID_full"] = df[name_col].astype(str)
    df["Enhancer_short"] = df["Enhancer_ID_full"].apply(short_name)

    # Sort: intronic first, then upstream, then unknown; within each by H3K4me1 desc
    order_map = {"intronic": 0, "upstream": 1, "unknown": 2}
    df["_loc_order"] = df["Location"].map(lambda x: order_map.get(x, 2))
    df = df.sort_values(["_loc_order", "H3K4me1_mean"], ascending=[True, False]).drop(columns=["_loc_order"]).reset_index(drop=True)

    # Final column order
    out_cols = [
        "Enhancer_short",
        "Enhancer_ID_full",
        "Location",
        "H3K4me1_mean",
        "H3K27ac_mean",
        "H3K4me3_mean",
    ]
    df_out = df[out_cols].copy()

    wb = Workbook()
    ws = wb.active
    ws.title = "Fig2C_SupportingData"

    # ---- Title block ----
    ws["A1"] = args.title
    ws["A2"] = args.subtitle
    ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Source: {args.tsv}"

    title_font = Font(bold=True, size=16)
    subtitle_font = Font(italic=True, size=11, color="444444")
    meta_font = Font(size=10, color="666666")

    ws["A1"].font = title_font
    ws["A2"].font = subtitle_font
    ws["A3"].font = meta_font

    # Merge title rows across columns
    last_col_letter = get_column_letter(len(df_out.columns))
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws.merge_cells(f"A2:{last_col_letter}2")
    ws.merge_cells(f"A3:{last_col_letter}3")

    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 16

    # ---- Header row ----
    header_row = 5
    header_fill = PatternFill("solid", fgColor="1F4E79")  # dark blue
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for j, colname in enumerate(df_out.columns, start=1):
        cell = ws.cell(row=header_row, column=j, value=colname)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    ws.row_dimensions[header_row].height = 22

    # ---- Data rows ----
    start_row = header_row + 1
    for i, row in enumerate(df_out.itertuples(index=False), start=start_row):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    # Number formatting for signal columns
    # Columns: H3K4me1_mean (4), H3K27ac_mean (5), H3K4me3_mean (6)
    for r in range(start_row, start_row + len(df_out)):
        for c in [4, 5, 6]:
            ws.cell(row=r, column=c).number_format = "0.000"

    # Freeze panes (keeps title + header visible)
    ws.freeze_panes = f"A{start_row}"

    # Add Excel Table object
    end_row = start_row + len(df_out) - 1
    table_ref = f"A{header_row}:{last_col_letter}{end_row}"
    tab = Table(displayName="Fig2C_Signals", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Align text columns left
    for r in range(start_row, end_row + 1):
        ws.cell(r, 1).alignment = Alignment(horizontal="left")
        ws.cell(r, 2).alignment = Alignment(horizontal="left")
        ws.cell(r, 3).alignment = Alignment(horizontal="center")

    # Autosize
    autosize_columns(ws)

    # ---- Footnotes ----
    foot_start = end_row + 3
    ws[f"A{foot_start}"] = "Footnotes"
    ws[f"A{foot_start}"].font = Font(bold=True, size=12)

    notes = [
        "1) Genome build: mm9. Enhancer intervals and signal summaries are in mm9 coordinates.",
        "2) Signals represent mean bigWig values across each enhancer interval for WT hindlimb E11.5 (merged track).",
        "3) H3K4me1 and H3K27ac are enhancer-associated marks; H3K4me3 is a promoter-associated mark included as a control for promoter proximity/overlap.",
        "4) Location annotation indicates whether the enhancer lies within the Gli3 intronic region (intronic) or 5′ of the gene body (upstream), as defined in the Fig2 coordinate track set.",
        "5) Sorting: enhancers are grouped by Location (intronic first, then upstream), and within each group sorted by H3K4me1_mean (descending).",
        "6) This table supports Fig. 2C, which displays the same mean signals on a shared axis to enable direct mark-by-mark comparison per enhancer.",
        "7) If a value is 0.000, it indicates either negligible signal across the interval or absence of signal in the provided bigWig track at that region (not a missing-data code).",
        "8) File provenance: Source TSV path is recorded in row 3. Track files are in your Fig2 folder (raw/*.bw) and enhancer tracks (coords/*.bw).",
    ]

    for k, line in enumerate(notes, start=foot_start + 1):
        ws[f"A{k}"] = line
        ws[f"A{k}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")
        ws.merge_cells(f"A{k}:{last_col_letter}{k}")
        ws.row_dimensions[k].height = 28

    # Final tweaks
    ws.sheet_view.showGridLines = False

    wb.save(args.out)
    print(f"[OK] Wrote: {args.out}")

if __name__ == "__main__":
    main()
